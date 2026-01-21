from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def _safe_sheet_name(name: str) -> str:
    # Excel: max 31 chars, no []:*?/\
    name = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip()
    return (name[:31] if len(name) > 31 else name) or "Sheet"


def _style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")  # azul oscuro
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize_columns(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_df(ws, df: pd.DataFrame):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _style_worksheet(ws)
    _autosize_columns(ws)


def load_tipologias_from_dir(in_dir: str) -> pd.DataFrame:
    p = Path(in_dir)
    files = sorted(p.glob("*__tipologias.parquet"))
    if not files:
        raise FileNotFoundError(f"No se encontraron *__tipologias.parquet en: {in_dir}")

    dfs = [pd.read_parquet(f) for f in files]
    df = pd.concat(dfs, ignore_index=True)

    # derivadas
    df["precio_m2"] = df["precio_desde"] / df["area_m2"]
    return df


def dedupe_tipologias(df: pd.DataFrame) -> pd.DataFrame:
    dedupe_key = ["url", "modelo", "dormitorios", "banos", "area_m2", "piso_min", "piso_max"]
    if "scraped_at" in df.columns:
        df = df.sort_values("scraped_at", ascending=False)
    return df.drop_duplicates(subset=dedupe_key, keep="first").reset_index(drop=True)


def compute_metrics(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    # Por proyecto
    g = df.groupby(["proyecto"], dropna=False)

    def _agg(x: pd.DataFrame) -> pd.Series:
        return pd.Series({
            "n_rows": len(x),
            "n_parse_ok": int((x.get("parse_ok") == True).sum()) if "parse_ok" in x.columns else None,
            "n_parse_fail": int((x.get("parse_ok") == False).sum()) if "parse_ok" in x.columns else None,
            "sum_unidades_disponibles": pd.to_numeric(x.get("unidades_disponibles"), errors="coerce").sum(),
            "min_precio": pd.to_numeric(x.get("precio_desde"), errors="coerce").min(),
            "avg_precio": pd.to_numeric(x.get("precio_desde"), errors="coerce").mean(),
            "max_precio": pd.to_numeric(x.get("precio_desde"), errors="coerce").max(),
            "min_precio_m2": pd.to_numeric(x.get("precio_m2"), errors="coerce").min(),
            "avg_precio_m2": pd.to_numeric(x.get("precio_m2"), errors="coerce").mean(),
            "max_precio_m2": pd.to_numeric(x.get("precio_m2"), errors="coerce").max(),
        })

    df_proj = g.apply(_agg).reset_index()

    # Global
    df_global = pd.DataFrame([{
        "n_rows": len(df),
        "n_proyectos": df["proyecto"].nunique(dropna=True),
        "n_parse_ok": int((df.get("parse_ok") == True).sum()) if "parse_ok" in df.columns else None,
        "n_parse_fail": int((df.get("parse_ok") == False).sum()) if "parse_ok" in df.columns else None,
        "sum_unidades_disponibles": pd.to_numeric(df.get("unidades_disponibles"), errors="coerce").sum(),
        "min_precio": pd.to_numeric(df.get("precio_desde"), errors="coerce").min(),
        "avg_precio": pd.to_numeric(df.get("precio_desde"), errors="coerce").mean(),
        "max_precio": pd.to_numeric(df.get("precio_desde"), errors="coerce").max(),
        "min_precio_m2": pd.to_numeric(df.get("precio_m2"), errors="coerce").min(),
        "avg_precio_m2": pd.to_numeric(df.get("precio_m2"), errors="coerce").mean(),
        "max_precio_m2": pd.to_numeric(df.get("precio_m2"), errors="coerce").max(),
    }])

    return df_proj, df_global


def make_datetimes_naive(df: pd.DataFrame) -> pd.DataFrame:
    """
    Excel (openpyxl) no soporta datetimes con tzinfo.
    Convierte todas las columnas datetime tz-aware a tz-naive.
    """
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                # Si es tz-aware
                if hasattr(df[col].dt, "tz") and df[col].dt.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)
            except Exception:
                pass
    return df



def export_excel(in_dir: str, out_xlsx: str, per_project_sheets: bool = True) -> str:
    df_all = load_tipologias_from_dir(in_dir)
    df_all = dedupe_tipologias(df_all)
    df_all = make_datetimes_naive(df_all)

    df_proj, df_global = compute_metrics(df_all)
    df_proj = make_datetimes_naive(df_proj)
    df_global = make_datetimes_naive(df_global)

    wb = Workbook()
    wb.remove(wb.active)

    # Consolidado
    ws = wb.create_sheet(_safe_sheet_name("Consolidado"))
    _write_df(ws, df_all)

    # Métricas
    ws = wb.create_sheet(_safe_sheet_name("Métricas"))
    _write_df(ws, df_global)
    ws.append([])
    for r in dataframe_to_rows(df_proj, index=False, header=True):
        ws.append(r)
    _style_worksheet(ws)
    _autosize_columns(ws)

    # Una hoja por proyecto (opcional)
    if per_project_sheets:
        for proj, d in df_all.groupby("proyecto"):
            name = _safe_sheet_name(f"Proyecto - {proj}")
            d = make_datetimes_naive(d)

            ws = wb.create_sheet(name)
            _write_df(ws, d.reset_index(drop=True))

    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx

